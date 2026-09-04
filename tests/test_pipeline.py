"""
Tests de pipeline (S15) y del módulo compartido de agrupación (S8/S11).

Pipeline REAL: PDF chico creado con PyMuPDF -> pdf_loader.cargar_paginas
(carga perezosa, sin "imagen" materializada) -> extraer_planilla_pdf -> el
generador de Excel -> abrir el .xlsx con openpyxl y verificar hojas y dedupe.
Lo ÚNICO mockeado es la llamada HTTP a Gemini (gemini_extractor._cliente).

También cubren S8 (agrupar_por_curso es la fuente única de verdad de GUI y
Excel) y S11 (dedupe de estudiantes por `no` y por nombre).

Uso (desde la raíz del proyecto):
    python -m unittest discover -s tests -v
"""

import json
import os
import sys
import tempfile
import unittest
from unittest import mock

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import pymupdf as fitz  # noqa: E402
from PIL import Image  # noqa: E402

from excel.agrupacion import agrupar_por_curso, combinar_estudiantes  # noqa: E402
from pdf_processing import pdf_loader  # noqa: E402
from vision import gemini_extractor as gem  # noqa: E402


# ---------------------------------------------------------------------- #
# Fixtures del pipeline
# ---------------------------------------------------------------------- #

def _json_planilla(grupo: str, estudiantes: list) -> str:
    """JSON en el formato crudo que devuelve el modelo (encabezado + alumnos)."""
    planilla = {
        "encabezado": {
            "institucion": "INSTITUCION PRUEBA",
            "sede": "SEDE PRINCIPAL",
            "año_lectivo": "2026",
            "jornada": "MAÑANA",
            "grupo": grupo,
            "asignatura": "MATEMATICAS",
            "docente": "DOCENTE PRUEBA",
            "periodo": 3,
        },
        "estudiantes": estudiantes,
    }
    return json.dumps(planilla, ensure_ascii=False)


class _FakeResponse:
    def __init__(self, texto):
        self.text = texto


class _FakeModels:
    """Devuelve una planilla distinta por llamada (una por página del PDF)."""

    def __init__(self, respuestas_por_pagina: list):
        self._respuestas = respuestas_por_pagina
        self._contador = 0

    def generate_content(self, **kwargs):
        if self._contador >= len(self._respuestas):
            raise AssertionError("generate_content se llamó más veces que páginas")
        texto = self._respuestas[self._contador]
        self._contador += 1
        return _FakeResponse(texto)


class _FakeClient:
    def __init__(self, respuestas_por_pagina: list):
        self.models = _FakeModels(respuestas_por_pagina)


def _estudiante(no, nombre):
    return {
        "no": no, "nombre": nombre, "ev_anteriores": [45, 45],
        "area_trabajo": [40, 50], "retirado": False, "revisar": [False, False],
    }


def _crear_pdf_prueba(ruta, paginas=2, texto="Planilla de prueba"):
    """Crea un PDF real chico con PyMuPDF (texto por página)."""
    doc = fitz.open()
    for _ in range(paginas):
        page = doc.new_page()
        page.insert_text((72, 72), texto, fontname="helv", fontsize=12)
    doc.save(ruta)
    doc.close()


# ---------------------------------------------------------------------- #
# S8: agrupar_por_curso
# ---------------------------------------------------------------------- #

class TestAgruparPorCurso(unittest.TestCase):

    def _planilla(self, grupo):
        return {"encabezado": {"grupo": grupo}, "estudiantes": []}

    def test_orden_de_primera_aparicion(self):
        ps = [self._planilla("0302"), self._planilla("0201"), self._planilla("0302")]
        por_curso, orden = agrupar_por_curso(ps)
        # Las claves ahora son tuplas (grupo, asignatura); sin asignatura, el
        # fallback es "SIN ASIGNATURA".
        self.assertEqual(orden, [("0302", "SIN ASIGNATURA"), ("0201", "SIN ASIGNATURA")])
        self.assertEqual(len(por_curso[("0302", "SIN ASIGNATURA")]), 2)
        self.assertEqual(len(por_curso[("0201", "SIN ASIGNATURA")]), 1)

    def test_grupo_vacio_o_ausente_va_a_sin_curso(self):
        ps = [
            {"encabezado": {"grupo": ""}, "estudiantes": []},
            {"encabezado": {}, "estudiantes": []},
        ]
        por_curso, orden = agrupar_por_curso(ps)
        self.assertEqual(orden, [("SIN CURSO", "SIN ASIGNATURA")])
        self.assertEqual(len(por_curso[("SIN CURSO", "SIN ASIGNATURA")]), 2)

    def test_no_muta_las_planillas(self):
        ps = [self._planilla("0302")]
        agrupar_por_curso(ps)
        self.assertEqual(ps, [self._planilla("0302")])


# ---------------------------------------------------------------------- #
# S11: combinar_estudiantes
# ---------------------------------------------------------------------- #

class TestCombinarEstudiantes(unittest.TestCase):

    def test_concatena_paginas_en_orden(self):
        p1 = {"estudiantes": [_estudiante(1, "GARCIA LUZ"), _estudiante(2, "PEREZ JUAN")]}
        p2 = {"estudiantes": [_estudiante(3, "RODRIGUEZ ANA")]}
        res = combinar_estudiantes([p1, p2])
        self.assertEqual([e["no"] for e in res], [1, 2, 3])

    def test_dedupe_por_no_entre_paginas_primer_gana(self):
        p1 = {"estudiantes": [_estudiante(2, "PEREZ JUAN"), _estudiante(5, "LOPEZ SOL")]}
        p2 = {"estudiantes": [_estudiante(2, "PEREZ JUAN (solapado)"), _estudiante(7, "DIAZ MAR")]}
        res = combinar_estudiantes([p1, p2])
        self.assertEqual([e["no"] for e in res], [2, 5, 7])
        # El primero gana: los valores del que estaba en la primera página.
        self.assertEqual(res[0]["nombre"], "PEREZ JUAN")

    def test_dedupe_por_no_ignora_nombres_distintos(self):
        # Mismo "no" con nombre distinto sigue siendo un duplicado (el número
        # de lista identifica al alumno): el primero gana.
        p1 = {"estudiantes": [_estudiante(4, "MORA RITA")]}
        p2 = {"estudiantes": [_estudiante(4, "MORA RITA ALIAS")]}
        res = combinar_estudiantes([p1, p2])
        self.assertEqual(len(res), 1)
        self.assertEqual(res[0]["nombre"], "MORA RITA")

    def test_dedupe_por_nombre_cuando_no_es_cero_o_ausente(self):
        # Sin número (0/ausente) no se puede deduplicar por "no": cae la regla
        # de nombre normalizado (mayúsculas y espacios colapsados).
        p1 = {"estudiantes": [{"no": 0, "nombre": "PEREZ  JUAN", "area_trabajo": [1, 2]},
                              {"no": 0, "nombre": "GARCIA LUZ", "area_trabajo": [3, 4]}]}
        p2 = {"estudiantes": [{"no": 0, "nombre": " perez juan ", "area_trabajo": [9, 9]}]}
        res = combinar_estudiantes([p1, p2])
        self.assertEqual(len(res), 2)
        self.assertEqual([e["nombre"] for e in res], ["PEREZ  JUAN", "GARCIA LUZ"])
        self.assertEqual(res[0]["area_trabajo"], [1, 2])  # el primero gana

    def test_no_0_con_nombres_distintos_se_conservan(self):
        p1 = {"estudiantes": [{"no": 0, "nombre": "ALFA", "area_trabajo": [1, 1]},
                              {"no": 0, "nombre": "BETA", "area_trabajo": [2, 2]}]}
        res = combinar_estudiantes([p1])
        self.assertEqual(len(res), 2)

    def test_registros_no_y_nombre_no_se_cruzan(self):
        # Un alumno con no=5 y otro sin número con el MISMO nombre conviven:
        # la regla de dedupe por "no" y la de nombre aplican en casos distintos.
        p1 = {"estudiantes": [{"no": 0, "nombre": "ALMA LUZ", "area_trabajo": [1, 1]}]}
        p2 = {"estudiantes": [{"no": 5, "nombre": "ALMA LUZ", "area_trabajo": [2, 2]}]}
        res = combinar_estudiantes([p1, p2])
        self.assertEqual(len(res), 2)

    def test_entradas_sin_estudiantes_y_none(self):
        self.assertEqual(combinar_estudiantes([]), [])
        self.assertEqual(combinar_estudiantes([{"estudiantes": None}]), [])
        self.assertEqual(combinar_estudiantes([{}, {"estudiantes": []}]), [])


# ---------------------------------------------------------------------- #
# S15: pipeline real (solo la llamada HTTP a Gemini está mockeada)
# ---------------------------------------------------------------------- #

class TestPipelineS15(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.dir_tmp = tempfile.mkdtemp(prefix="notas_pipeline_")
        cls.ruta_pdf = os.path.join(cls.dir_tmp, "planillas_prueba.pdf")
        _crear_pdf_prueba(cls.ruta_pdf, paginas=2)

    @classmethod
    def tearDownClass(cls):
        import shutil
        shutil.rmtree(cls.dir_tmp, ignore_errors=True)

    def test_cargar_paginas_lazy_sin_imagen_y_render_pil(self):
        paginas = pdf_loader.cargar_paginas(self.ruta_pdf, dpi=150)
        self.assertEqual(len(paginas), 2)
        try:
            for pag in paginas:
                # S7: NUNCA se materializa "imagen" al cargar.
                self.assertNotIn("imagen", pag, "cargar_paginas no debe materializar imágenes")
                self.assertIn("indice", pag)
                self.assertGreater(pag["ancho_px"], 0)
                self.assertGreater(pag["alto_px"], 0)
                self.assertTrue(callable(pag["_render"]))
                self.assertIsNotNone(pag["_doc"])
            # Render perezoso: recién acá se rasteriza la página.
            for pag in paginas:
                img = pag["_render"]()
                self.assertIsInstance(img, Image.Image)
                self.assertEqual(img.mode, "RGB")
                img.close()
        finally:
            pdf_loader.cerrar_paginas(paginas)

    def test_circuito_completo_pdf_vision_excel_sin_duplicados(self):
        # Página 1 -> curso 0201 (2 alumnos). Página 2 -> curso 0302 (2 alumnos
        # + un SOLAPADO: el alumno nº 2 repetido, escaneo duplicado; S11 debe
        # descartar la copia en la GUI y en el Excel).
        respuestas = [
            _json_planilla("0201", [_estudiante(1, "GARCIA LUZ"), _estudiante(2, "PEREZ JUAN")]),
            _json_planilla("0302", [
                _estudiante(1, "RODRIGUEZ ANA"),
                _estudiante(2, "PEREZ JUAN"),
                _estudiante(2, "PEREZ JUAN"),  # duplicado (solapado)
            ]),
        ]

        paginas = pdf_loader.cargar_paginas(self.ruta_pdf, dpi=150)
        try:
            # Lo ÚNICO mockeado: la conexión HTTP a Gemini.
            with mock.patch.object(gem, "_cliente", return_value=_FakeClient(respuestas)):
                planillas, fallidas = gem.extraer_planilla_pdf(
                    paginas, api_key="clave-de-prueba"
                )
            self.assertEqual(fallidas, [])
            self.assertEqual(len(planillas), 2)
            self.assertEqual(planillas[0]["encabezado"]["grupo"], "0201")
            self.assertEqual(planillas[1]["encabezado"]["grupo"], "0302")
            self.assertEqual(len(planillas[1]["estudiantes"]), 3)

            # El circuito termina en el Excel real.
            ruta_xlsx = os.path.join(self.dir_tmp, "notas_pipeline.xlsx")
            from excel.generar_excel_notas import generar_excel_asignatura
            generar_excel_asignatura(planillas, ruta_xlsx)

            import openpyxl
            wb = openpyxl.load_workbook(ruta_xlsx)
            self.assertEqual(
                wb.sheetnames,
                ["Curso 0201 - MATEMATICAS", "Curso 0302 - MATEMATICAS"],
            )

            ws = wb["Curso 0302 - MATEMATICAS"]
            # Filas de alumnos = col A (no) y col B (nombre) pobladas; la fila
            # del legendario "Amarillo = ..." tiene col A vacía y no cuenta.
            nombres = [
                fila[1].value for fila in ws.iter_rows(min_row=11)
                if fila[0].value is not None and fila[1].value is not None
            ]
            self.assertEqual(len(nombres), 2, f"El duplicado entró al Excel: {nombres}")
            self.assertEqual(nombres.count("PEREZ JUAN"), 1, "El duplicado aparece 2 veces")
        finally:
            pdf_loader.cerrar_paginas(paginas)

    def test_cerrar_paginas_idempotente(self):
        paginas = pdf_loader.cargar_paginas(self.ruta_pdf, dpi=150)
        pdf_loader.cerrar_paginas(paginas)
        # Segunda llamada: no debe explotar (doc ya cerrado/marcado).
        pdf_loader.cerrar_paginas(paginas)
        pdf_loader.cerrar_paginas(None)
        pdf_loader.cerrar_paginas([])
        # El dict quedó marcado como cerrado.
        self.assertIsNone(paginas[0]["_doc"])


if __name__ == "__main__":
    unittest.main()