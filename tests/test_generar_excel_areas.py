"""
Test del ancho dinámico de columnas de área de trabajo en el generador de Excel.

Cubre el camino de salida del pipeline (spec v2): la extracción puede devolver
de 1 a 16 notas por alumno y el generador debe reflejar ESE ancho real en el
Excel, no un fijo de 2 columnas. También cubre la inferencia cuando no se
declara n_area_trabajo pero sí hay notas observadas, y la regla de fallback a
la planilla tradicional de 2 notas cuando no hay nada que medir.

Uso (desde la raíz del proyecto):
    python -m tests.test_generar_excel_areas
"""

import os
import sys
import tempfile
import unittest
import copy

# La raíz del proyecto se agrega al path para poder importar el paquete excel/.
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

import excel.generar_excel_notas as generador  # noqa: E402
from excel.agrupacion import agrupar_por_curso, combinar_estudiantes, _asignatura_limpia  # noqa: E402


def _planilla(area_por_alumno, n_area_trabajo=None, periodo=3):
    """Construye una planilla con `area_por_alumno` (lista de listas o listas
    de notas por alumno). Quien dé None en `area_por_alumno` es un retirado."""
    encabezado = {
        "institucion": "INSTITUCION DEMO",
        "sede": "SEDE",
        "año_lectivo": "2026",
        "jornada": "MAÑANA",
        "grupo": "0302",
        "asignatura": "MATEMATICAS",
        "docente": "DOCENTE DEMO",
        "periodo": periodo,
    }
    if n_area_trabajo is not None:
        encabezado["n_area_trabajo"] = n_area_trabajo

    estudiantes = []
    for i, notas in enumerate(area_por_alumno, start=1):
        retirado = notas is None
        estudiantes.append({
            "no": i,
            "nombre": f"ALUMNO {i}",
            "ev_anteriores": [45, 45] if periodo > 1 else [],
            "area_trabajo": notas if not retirado else None,
            "retirado": retirado,
            "revisar": [False] * (len(notas) if not retirado else 0),
        })
    return {"encabezado": encabezado, "estudiantes": estudiantes}


def _hoja_cargada(planilla):
    """Genera el Excel de la planilla y devuelve la hoja activa ya cargada."""
    directorio = tempfile.mkdtemp(prefix="notas_areas_")
    ruta = os.path.join(directorio, "salida.xlsx")
    generador.generar_excel_planilla(planilla, ruta)
    wb = openpyxl.load_workbook(ruta)
    return wb[wb.sheetnames[0]]


class TestAnchoDinamicoAreas(unittest.TestCase):
    """El ancho de las columnas de área sigue la cantidad real de notas."""

    def test_planilla_de_5_notas_escribe_5_columnas_y_formula_EI(self):
        # Periodo 3 => 2 columnas de definitiva previa (C y D), área arranca en E.
        planilla = _planilla(
            [
                [40, 50, 60, 70, 80],
                [41, 51, 61, 71, 81],
                [42, 52, 62, 72, 82],
            ],
            n_area_trabajo=5,
        )
        ws = _hoja_cargada(planilla)

        # Fila 10 = header (8 filas informativas + 2). Verifico los títulos.
        headers = [ws.cell(row=10, column=j).value for j in range(1, 12)]
        for k in range(1, 6):
            self.assertIn(f"Área Trabajo {k}", headers)
        self.assertIn("Definitiva Periodo 3", headers)

        # Alumno 1 (fila 11): nota 5 en la columna I (9), nota 1 en E (5),
        # y la definitiva en J (10) con SUM(E:I)/5 (divide por las 5 columnas).
        self.assertEqual(ws.cell(row=11, column=5).value, 40)
        self.assertEqual(ws.cell(row=11, column=9).value, 80)
        # La fórmula usa el rango completo de las 5 celdas de área (E:I)
        # dividido por el total de columnas, no por lo que trae el alumno.
        self.assertEqual(
            ws.cell(row=11, column=10).value,
            '=IFERROR(SUM(E11:I11)/5,"")',
        )

    def test_alumno_con_menos_notas_deja_celda_vacia(self):
        # Un alumno con 4 notas en una planilla de 5: la 5ª celda (columna I)
        # queda None (vacía), sin excepción por índice fuera de rango.
        planilla = _planilla(
            [
                [40, 50, 60, 70, 80],
                [41, 51, 61, 71],  # solo 4 notas
            ],
            n_area_trabajo=5,
        )
        ws = _hoja_cargada(planilla)
        self.assertEqual(ws.cell(row=12, column=5).value, 41)
        self.assertEqual(ws.cell(row=12, column=8).value, 71)  # 4ª nota (H)
        self.assertIsNone(ws.cell(row=12, column=9).value)  # 5ª celda vacía (I)

    def test_dos_notas_de_cinco_divide_por_el_total_de_columnas(self):
        # Un alumno con 2 notas en una planilla de 5: la definitiva divide por
        # 5 (el total de columnas de área), no por las 2 notas presentes: las
        # celdas en blanco son actividades no realizadas (cuentan como 0).
        planilla = _planilla(
            [
                [40, 50, 60, 70, 80],
                [41, 51],  # solo 2 notas de 5
            ],
            n_area_trabajo=5,
        )
        ws = _hoja_cargada(planilla)
        self.assertEqual(ws.cell(row=12, column=5).value, 41)
        self.assertEqual(ws.cell(row=12, column=6).value, 51)
        self.assertIsNone(ws.cell(row=12, column=7).value)  # 3ª celda vacía (G)
        self.assertEqual(
            ws.cell(row=12, column=10).value,
            '=IFERROR(SUM(E12:I12)/5,"")',
        )

    def test_una_sola_nota_genera_una_columna_y_formula_EE(self):
        # Periodo 3 => área arranca en E; una sola nota => col_def = F, rango E:E.
        planilla = _planilla(
            [[55]],
            n_area_trabajo=1,
        )
        ws = _hoja_cargada(planilla)

        headers = [ws.cell(row=10, column=j).value for j in range(1, 8)]
        self.assertIn("Área Trabajo 1", headers)
        self.assertNotIn("Área Trabajo 2", headers)

        self.assertEqual(ws.cell(row=11, column=5).value, 55)
        self.assertEqual(
            ws.cell(row=11, column=6).value,
            '=IFERROR(SUM(E11:E11)/1,"")',
        )

    def test_inferencia_sin_declaracion_usa_lo_observado(self):
        # Sin n_area_trabajo en el encabezado pero con 3 notas por alumno:
        # el ancho se infiere de lo observado (3 columnas).
        planilla = _planilla(
            [
                [40, 50, 60],
                [41, 51, 61],
            ],
            n_area_trabajo=None,
        )
        ws = _hoja_cargada(planilla)

        headers = [ws.cell(row=10, column=j).value for j in range(1, 9)]
        self.assertIn("Área Trabajo 1", headers)
        self.assertIn("Área Trabajo 3", headers)
        self.assertNotIn("Área Trabajo 4", headers)

        # Área arranca en E (3), con 3 notas => col_def = 5 + 3 = 8 (H);
        # rango E:G en la celda de definitiva.
        self.assertEqual(ws.cell(row=11, column=8).value,
                         '=IFERROR(SUM(E11:G11)/3,"")')


class TestCalcularNAreas(unittest.TestCase):
    """calcular_n_areas: el ancho de la zona de área de trabajo (1 a 16 notas)."""

    def _enc_y_estudiantes(self, notas_por_alumno, n_area_trabajo):
        planilla = _planilla(notas_por_alumno, n_area_trabajo=n_area_trabajo)
        return planilla["encabezado"], planilla["estudiantes"]

    def test_declarado_int_5(self):
        enc, est = self._enc_y_estudiantes([[40, 50]], n_area_trabajo=5)
        self.assertEqual(generador.calcular_n_areas(enc, est), 5)

    def test_sin_declarado_con_3_notas_observadas(self):
        enc, est = self._enc_y_estudiantes(
            [[40, 50, 60], [41, 51, 61]], n_area_trabajo=None
        )
        self.assertEqual(generador.calcular_n_areas(enc, est), 3)

    def test_declarado_1_con_observado_4_prevalece_observado(self):
        enc, est = self._enc_y_estudiantes([[40, 50, 60, 70]], n_area_trabajo=1)
        self.assertEqual(generador.calcular_n_areas(enc, est), 4)

    def test_declarado_bool_true_ignorado(self):
        # Un bool no es un conteo: se ignora y manda lo observado.
        enc, est = self._enc_y_estudiantes([[40, 50, 60]], n_area_trabajo=True)
        self.assertEqual(generador.calcular_n_areas(enc, est), 3)

    def test_sin_declarado_ni_notas_fallback_2(self):
        enc, est = self._enc_y_estudiantes([[]], n_area_trabajo=None)
        self.assertEqual(generador.calcular_n_areas(enc, est), 2)

    def test_declarado_16(self):
        enc, est = self._enc_y_estudiantes([[40] * 16], n_area_trabajo=16)
        self.assertEqual(generador.calcular_n_areas(enc, est), 16)

    def test_declarado_17_ignorado_y_observado_2(self):
        # 17 está fuera del rango 1..16: se ignora (0) y manda lo observado (2).
        enc, est = self._enc_y_estudiantes([[40, 50]], n_area_trabajo=17)
        self.assertEqual(generador.calcular_n_areas(enc, est), 2)


# ---------------------------------------------------------------------- #
# Diagnóstico real: un mismo curso puede tener asignaturas DISTINTAS en el
# mismo PDF. La agrupación debe separar por (asignatura, curso) para no
# descartar la segunda página de cada curso (notas 0501-Naturales/0502-Inf.)
# ---------------------------------------------------------------------- #

def _diagnostico_planilla(grupo, asignatura, area_por_alumno):
    """Construye una planilla de diagnóstico con `area_por_alumno` (notas por
    alumno, una lista por alumno) en el formato del generador."""
    estudiantes = []
    for i, notas in enumerate(area_por_alumno, start=1):
        estudiantes.append({
            "no": i,
            "nombre": f"ALUMNO {i}",
            "ev_anteriores": [45, 45],
            "area_trabajo": list(notas),
            "retirado": False,
            "revisar": [False] * len(notas),
        })
    return {
        "encabezado": {
            "institucion": "INST", "sede": "SEDE", "año_lectivo": "2026",
            "jornada": "M", "grupo": grupo, "asignatura": asignatura,
            "docente": "DOC", "periodo": 3,
        },
        "estudiantes": estudiantes,
    }


def _planillas_diagnostico():
    """Las 4 planillas del caso diagnosticado (2 cursos x 2 asignaturas).

    - Pág 1: 0501 / TECNOLOGIA E INFORMATICA
    - Pág 2: 0501 / CIENCIAS NATURALES Y ED. AMBIENTAL (4 notas por alumno)
    - Pág 3: 0502 / "0500101 CIENCIAS NATURALES Y ED. AMBIENTAL" (prefijo)
    - Pág 4: 0502 / TECNOLOGIA E INFORMATICA
    """
    return [
        _diagnostico_planilla("0501", "TECNOLOGIA E INFORMATICA",
                              [[40, 50], [41, 51]]),
        _diagnostico_planilla("0501", "CIENCIAS NATURALES Y ED. AMBIENTAL",
                              [[40, 50, 60, 70], [41, 51, 61, 71]]),
        _diagnostico_planilla("0502", "0500101 CIENCIAS NATURALES Y ED. AMBIENTAL",
                              [[30, 40], [31, 41], [32, 42]]),
        _diagnostico_planilla("0502", "TECNOLOGIA E INFORMATICA",
                              [[20, 30], [21, 31]]),
    ]


class TestAgrupacionAsignaturaCurso(unittest.TestCase):
    """La clave de agrupación es (asignatura normalizada, curso)."""

    def test_cuatro_planillas_dan_cuatro_claves_distintas(self):
        por_curso, orden = agrupar_por_curso(_planillas_diagnostico())
        self.assertEqual(len(por_curso), 4)
        self.assertEqual(orden, [
            ("0501", "TECNOLOGIA E INFORMATICA"),
            ("0501", "CIENCIAS NATURALES Y ED. AMBIENTAL"),
            ("0502", "CIENCIAS NATURALES Y ED. AMBIENTAL"),  # prefijo removido
            ("0502", "TECNOLOGIA E INFORMATICA"),
        ])
        # Cada clave tiene exactamente 1 página (no se mezclan asignaturas).
        for clave, paginas in por_curso.items():
            self.assertEqual(len(paginas), 1, f"La clave {clave} no debería mezclar páginas")

    def test_dos_paginas_iguales_de_misma_asignatura_curso_una_sola_clave(self):
        # Misma planilla 0501-Naturales escaneada 2 veces: 1 sola clave y los
        # alumnos sin duplicados (dedupe S11 intacto DENTRO de la asignatura).
        p = _diagnostico_planilla("0501", "CIENCIAS NATURALES Y ED. AMBIENTAL",
                                  [[40, 50, 60], [41, 51, 61]])
        ps = [p, copy.deepcopy(p)]
        por_curso, orden = agrupar_por_curso(ps)
        self.assertEqual(len(por_curso), 1)
        self.assertEqual(len(orden), 1)
        paginas = por_curso[orden[0]]
        estudiantes = combinar_estudiantes(paginas)
        self.assertEqual(len(estudiantes), 2)
        self.assertEqual([e["no"] for e in estudiantes], [1, 2])


class TestAsignaturaLimpia(unittest.TestCase):
    """_asignatura_limpia: quita el código numérico y normaliza el nombre."""

    def test_quita_prefijo_de_codigo_numerico(self):
        self.assertEqual(
            _asignatura_limpia("0500101 CIENCIAS NATURALES Y ED. AMBIENTAL"),
            "CIENCIAS NATURALES Y ED. AMBIENTAL",
        )

    def test_sin_prefijo_queda_igual(self):
        self.assertEqual(
            _asignatura_limpia("TECNOLOGIA E INFORMATICA"),
            "TECNOLOGIA E INFORMATICA",
        )

    def test_colapsa_espacios_y_mayusculas(self):
        self.assertEqual(
            _asignatura_limpia("  tecnologia   e  informatica "),
            "TECNOLOGIA E INFORMATICA",
        )


class TestGenerarExcelDiagnostico(unittest.TestCase):
    """generar_excel_asignatura con las 4 planillas: 4 hojas, sin mezclar ni
    truncar notas (el diagnóstico perdía la segunda página de cada curso)."""

    def test_cuatro_hojas_por_asignatura_y_curso_sin_mezclar_ni_truncar(self):
        dir_tmp = tempfile.mkdtemp(prefix="notas_diag_")
        ruta = os.path.join(dir_tmp, "diag.xlsx")
        try:
            generador.generar_excel_asignatura(_planillas_diagnostico(), ruta)
            wb = openpyxl.load_workbook(ruta)

            self.assertEqual(len(wb.sheetnames), 4)

            def hoja_con(*fragmentos):
                return [
                    s for s in wb.sheetnames
                    if all(f in s for f in fragmentos)
                ]

            # Una hoja para cada (curso, asignatura).
            self.assertEqual(len(hoja_con("0501", "TECNOLOGIA")), 1)
            self.assertEqual(len(hoja_con("0501", "CIENCIAS")), 1)
            self.assertEqual(len(hoja_con("0502", "CIENCIAS")), 1)
            self.assertEqual(len(hoja_con("0502", "TECNOLOGIA")), 1)

            # La hoja 0501-CIENCIAS lleva SUS alumnos con SUS notas: el alumno
            # 2 tiene 4 notas (no se trunca a 2), y no aparece el de otra
            # asignatura/curso.
            ws_ciencias = wb[hoja_con("0501", "CIENCIAS")[0]]
            encabezados = [
                ws_ciencias.cell(row=10, column=j).value for j in range(1, 12)
            ]
            self.assertIn("Área Trabajo 4", encabezados)
            self.assertNotIn("Área Trabajo 5", encabezados)
            # Fila 11 = alumno 1, fila 12 = alumno 2; la 4ª nota del alumno 2
            # (área arranca en E, columna 8 = nota 4) debe conservarse.
            self.assertEqual(ws_ciencias.cell(row=12, column=8).value, 71)
            # No se filtraron las notas de otra asignatura: 2 alumnos nada más.
            nombres = [
                fila[1].value for fila in ws_ciencias.iter_rows(min_row=11)
                if fila[0].value is not None and fila[1].value is not None
            ]
            self.assertEqual(len(nombres), 2)
            self.assertNotIn("ALUMNO 3", nombres)
        finally:
            import shutil
            shutil.rmtree(dir_tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()