"""
Test del generador de Excel.

Genera un Excel de demo llamando directamente a `generar_excel_asignatura`
(no con runpy, para evitar el RuntimeWarning de importar bajo __main__) y
valida que el .xlsx existe y tiene las tres hojas de curso esperadas
(Curso 0302, 0401 y 0501), además de fórmulas de definitiva.

La salida se escribe en una carpeta temporal (tempfile.mkdtemp), nunca en una
ruta fija del usuario (S5).

Uso (desde la raíz del proyecto):
    python -m tests.test_generar_excel
"""

import os
import sys
import tempfile
import unittest

# La raíz del proyecto se agrega al path para poder importar el paquete excel/.
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import excel.generar_excel_notas  # noqa: E402,F401

_HOJAS_ESPERADAS = [
    "Curso 0302 - MATEMATICAS",
    "Curso 0401 - MATEMATICAS",
    "Curso 0501 - MATEMATICAS",
]


def _planilla(grupo: str, periodo: int) -> dict:
    """Construye una planilla mínima en el formato que espera el generador."""
    return {
        "encabezado": {
            "institucion": "INSTITUCION DEMO",
            "sede": "SEDE",
            "año_lectivo": "2026",
            "jornada": "MAÑANA",
            "grupo": grupo,
            "asignatura": "MATEMATICAS",
            "docente": "DOCENTE DEMO",
            "periodo": periodo,
        },
        "estudiantes": [
            {
                "no": 1, "nombre": "ALUMNO UNO", "ev_anteriores": [45, 45],
                "area_trabajo": [40, 50], "retirado": False,
                "revisar": [False, False],
            },
            {
                "no": 2, "nombre": "ALUMNO DOS", "ev_anteriores": [45],
                "area_trabajo": None, "retirado": True, "revisar": [False, False],
            },
        ],
    }


def _demo_planillas() -> list:
    """Simula el caso real: un PDF con 3 cursos de la misma asignatura."""
    return [_planilla("0302", 3), _planilla("0401", 3), _planilla("0501", 3)]


class TestGeneradorExcel(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # Carpeta temporal: portable, sin rutas fijas a C:\Users\JUAN (S5).
        cls.dir_salida = tempfile.mkdtemp(prefix="notas_test_")
        cls.ruta_xlsx = os.path.join(cls.dir_salida, "notas_demo.xlsx")
        excel.generar_excel_notas.generar_excel_asignatura(
            _demo_planillas(), cls.ruta_xlsx
        )

    def test_archivo_generado_existe(self):
        self.assertTrue(
            os.path.exists(self.ruta_xlsx),
            f"El demo no generó el archivo en {self.ruta_xlsx}",
        )

    def test_hojas_de_curso_esperadas(self):
        self.assertTrue(os.path.exists(self.ruta_xlsx), "Falta el archivo generado")
        import openpyxl

        wb = openpyxl.load_workbook(self.ruta_xlsx)
        for hoja in _HOJAS_ESPERADAS:
            self.assertIn(
                hoja, wb.sheetnames,
                f"El Excel no tiene la hoja {hoja} (hojas: {wb.sheetnames})",
            )

    def test_existencia_formula_definitiva(self):
        self.assertTrue(os.path.exists(self.ruta_xlsx), "Falta el archivo generado")
        import openpyxl

        wb = openpyxl.load_workbook(self.ruta_xlsx)
        ws = wb["Curso 0302 - MATEMATICAS"]
        hay_formula = any(
            isinstance(cell.value, str) and cell.value.startswith("=IFERROR(AVERAGE")
            for row in ws.iter_rows()
            for cell in row
        )
        self.assertTrue(hay_formula, "No se encontró ninguna fórmula de definitiva en Curso 0302 - MATEMATICAS")


class TestGuardPeriodo(unittest.TestCase):
    """W-A: el generador debe abortar con ValueError ante un periodo inválido
    en lugar de producir un Excel corrupto (que pisa el nombre del alumno)."""

    def _dir(self):
        return tempfile.mkdtemp(prefix="notas_periodo_")

    def _pl(self, periodo):
        pl = _planilla("0302", 3)
        pl["encabezado"]["periodo"] = periodo
        return pl

    def test_periodo_cero_lanza_valorerror_asignatura(self):
        with self.assertRaises(ValueError) as cm:
            excel.generar_excel_notas.generar_excel_asignatura(
                [self._pl(0)], os.path.join(self._dir(), "a.xlsx")
            )
        self.assertIn("debe ser 1, 2, 3 o 4", str(cm.exception))

    def test_periodo_cinco_lanza_valorerror_asignatura(self):
        with self.assertRaises(ValueError) as cm:
            excel.generar_excel_notas.generar_excel_asignatura(
                [self._pl(5)], os.path.join(self._dir(), "a.xlsx")
            )
        self.assertIn("debe ser 1, 2, 3 o 4", str(cm.exception))

    def test_periodo_invalido_no_crea_archivo_asignatura(self):
        ruta = os.path.join(self._dir(), "no_debe_existir.xlsx")
        with self.assertRaises(ValueError):
            excel.generar_excel_notas.generar_excel_asignatura([self._pl(0)], ruta)
        self.assertFalse(os.path.exists(ruta), "No debería haberse creado el archivo")

    def test_periodo_invalido_lanza_valorerror_planilla(self):
        with self.assertRaises(ValueError):
            excel.generar_excel_notas.generar_excel_planilla(
                self._pl(0), os.path.join(self._dir(), "b.xlsx")
            )

    def test_periodos_validos_1_a_4_ok_asignatura(self):
        for p in (1, 2, 3, 4):
            ruta = os.path.join(self._dir(), f"ok_{p}.xlsx")
            excel.generar_excel_notas.generar_excel_asignatura([self._pl(p)], ruta)
            self.assertTrue(os.path.exists(ruta), f"Fallo el periodo válido {p}")

    def test_mensaje_valorerror_nombra_el_curso(self):
        with self.assertRaises(ValueError) as cm:
            excel.generar_excel_notas.generar_excel_asignatura(
                [self._pl(0)], os.path.join(self._dir(), "c.xlsx")
            )
        self.assertIn("0302", str(cm.exception))


if __name__ == "__main__":
    unittest.main()
