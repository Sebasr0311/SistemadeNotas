"""
Generador de Excel de notas a partir de datos extraídos de una planilla.

Esta es la segunda mitad del pipeline: recibe los datos YA EXTRAÍDOS
(en producción, extraídos por el modelo de visión celda por celda desde
la imagen de la planilla) y arma el Excel con fórmulas reales de Excel
(no valores fijos), para que la definitiva se recalcule sola si se
corrige una nota a mano en el Excel.

Uso: se llama una vez por cada planilla (cada hoja = un grupo+asignatura+periodo).
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def calcular_n_areas(enc, estudiantes) -> int:
    """Cantidad de columnas de área de trabajo (spec v2: de 1 a 16 notas).

    Misma lógica que usa el generador para dimensionar la hoja, extraída como
    función pura para que la pantalla de revisión de la GUI calcule EXACTAMENTE
    el mismo ancho (S8: lo que se muestra es lo que se escribe):
    - Si el encabezado declara n_area_trabajo como int (no bool) en 1..16,
      prevalece el declarado (celdas vacías al final si hay menos observadas).
    - Si no, se usa lo observado: la mayor cantidad de notas leídas por alumno.
    - Sin nada que medir, se cae al comportamiento histórico: planilla
      tradicional de 2 notas.
    """
    declarado_raw = enc.get("n_area_trabajo")
    declarado = (
        declarado_raw
        if isinstance(declarado_raw, int) and not isinstance(declarado_raw, bool)
        and 1 <= declarado_raw <= 16
        else 0
    )
    longitudes = [
        len(est.get("area_trabajo") or [])
        for est in estudiantes
    ]
    observado = max(longitudes) if longitudes else 0
    n_areas = max(declarado, observado)
    if n_areas == 0:
        # Sin declaración y sin ninguna nota: planillas tradicionales de 2 notas.
        n_areas = 2
    return n_areas


def _escribir_hoja(ws, planilla: dict):
    enc = planilla["encabezado"]
    estudiantes = planilla["estudiantes"]
    periodo = enc["periodo"]
    # Guard W1/W-A: nunca generar un Excel corrupto con un periodo inválido.
    # Si llega mal desde cualquier fuente (visión, edición manual, etc.), se
    # aborta en vez de que la nota pise la columna del nombre.
    if not (isinstance(periodo, int) and 1 <= periodo <= 4):
        raise ValueError(
            f"Periodo inválido ({periodo!r}) en el curso {enc.get('grupo', '?')}: "
            "debe ser 1, 2, 3 o 4."
        )
    n_ev_anteriores = periodo - 1

    # --- Ancho de la zona de área de trabajo (spec v2: 1 a 16 notas) ---
    n_areas = calcular_n_areas(enc, estudiantes)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    revisar_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # --- Encabezado informativo ---
    info_rows = [
        ("Institución", enc.get("institucion", "")),
        ("Sede", enc.get("sede", "")),
        ("Año lectivo", enc.get("año_lectivo", "")),
        ("Jornada", enc.get("jornada", "")),
        ("Grupo", enc.get("grupo", "")),
        ("Asignatura", enc.get("asignatura", "")),
        ("Docente", enc.get("docente", "")),
        ("Periodo", periodo),
    ]
    for i, (label, value) in enumerate(info_rows, start=1):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value)

    header_row = len(info_rows) + 2

    # --- Columnas dinámicas ---
    cols = ["No.", "Nombre del Alumno"]
    for p in range(1, n_ev_anteriores + 1):
        cols.append(f"Def. Periodo {p}")
    cols += [f"Área Trabajo {k}" for k in range(1, n_areas + 1)]
    cols.append(f"Definitiva Periodo {periodo}")
    incluir_anual = periodo == 4 and n_ev_anteriores == 3
    if incluir_anual:
        cols.append("Definitiva Anual")

    for j, title in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=j, value=title)
        c.font = bold
        c.fill = header_fill
        c.alignment = center
        c.border = border

    col_ev_start = 3
    col_at1 = col_ev_start + n_ev_anteriores
    col_def = col_at1 + n_areas
    col_anual = col_def + 1 if incluir_anual else None

    r = header_row + 1
    for est in estudiantes:
        ws.cell(row=r, column=1, value=est["no"]).border = border
        ws.cell(row=r, column=2, value=est["nombre"]).border = border

        if est.get("retirado"):
            ws.cell(row=r, column=2).font = Font(italic=True, color="999999")
            c = ws.cell(row=r, column=3, value="Retirado / sin datos")
            c.font = Font(italic=True, color="999999")
            r += 1
            continue

        ev = est.get("ev_anteriores") or []
        for k in range(n_ev_anteriores):
            v = ev[k] if k < len(ev) else None
            ws.cell(row=r, column=col_ev_start + k, value=v).border = border
            ws.cell(row=r, column=col_ev_start + k).alignment = center

        at = est.get("area_trabajo")
        revisar = est.get("revisar") or []
        for k in range(n_areas):
            v = at[k] if at and k < len(at) else None
            c = ws.cell(row=r, column=col_at1 + k, value=v)
            c.border = border
            c.alignment = center
            if k < len(revisar) and revisar[k]:
                c.fill = revisar_fill

        at1_ref = f"{get_column_letter(col_at1)}{r}"
        atN_ref = f"{get_column_letter(col_def - 1)}{r}"
        cdef = ws.cell(row=r, column=col_def)
        if at and len(at) >= 1:
            # La definitiva divide por el TOTAL de columnas de área (n_areas),
            # no por las notas que tiene el alumno: una celda en blanco es una
            # actividad no realizada y cuenta como 0 para el promedio del
            # periodo (con AVERAGE los vacíos se ignoraban y el promedio salía
            # sobre las notas presentes, inflando la definitiva).
            cdef.value = f"=IFERROR(SUM({at1_ref}:{atN_ref})/{n_areas},\"\")"
        cdef.border = border; cdef.alignment = center
        cdef.font = bold

        if incluir_anual:
            refs = [f"{get_column_letter(col_ev_start + k)}{r}" for k in range(n_ev_anteriores)]
            refs.append(f"{get_column_letter(col_def)}{r}")
            canual = ws.cell(row=r, column=col_anual)
            canual.value = f"=IFERROR(AVERAGE({','.join(refs)}),\"\")"
            canual.border = border; canual.alignment = center
            canual.font = bold

        r += 1

    # --- Leyenda ---
    r += 1
    leyenda = ws.cell(row=r, column=2, value="Amarillo = nota dudosa (tachón/letra ambigua), verificar contra la planilla física")
    leyenda.font = Font(italic=True, size=9, color="7F6000")
    ws.cell(row=r, column=2).fill = revisar_fill

    # --- Anchos de columna ---
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 34
    for j in range(3, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14


def generar_excel_planilla(planilla: dict, ruta_salida: str):
    """Genera un Excel de una sola planilla (un curso). Ver generar_excel_asignatura
    para el caso real de uso: un PDF con varias planillas de la misma asignatura."""
    wb = openpyxl.Workbook()
    ws = wb.active
    enc = planilla["encabezado"]
    ws.title = f"P{enc['periodo']} - {enc['grupo']}"
    _escribir_hoja(ws, planilla)
    wb.save(ruta_salida)
    return ruta_salida


def generar_excel_asignatura(planillas: list, ruta_salida: str):
    """
    Caso real de uso: un PDF sube TODAS las planillas de una misma asignatura
    (varios cursos/grupos). Se agrupan automáticamente por curso y se genera
    UN SOLO Excel con una hoja por cada curso.

    planillas: lista de dicts, cada uno con la misma forma que en
    generar_excel_planilla (un dict por planilla/página reconocida en el PDF).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # se reemplaza por una hoja por curso

    # S8: la agrupación por curso y la combinación de estudiantes viven en
    # excel/agrupacion.py, la MISMA fuente de verdad que usa la pantalla de
    # revisión de la GUI: el Excel escribe exactamente lo que la GUI muestra.
    # S11: combinar_estudiantes descarta estudiantes duplicados (páginas
    # repetidas o solapadas) antes de escribirlos.
    from excel.agrupacion import agrupar_por_curso, combinar_estudiantes

    por_curso, orden_grupos = agrupar_por_curso(planillas)

    nombres_usados = set()
    for clave in orden_grupos:
        # S8: cada clave es una tupla (grupo, asignatura): una hoja por
        # asignatura+curso, para que asignaturas distintas de un mismo curso
        # (caso diagnosticado) no se mezclen ni se pierdan notas.
        grupo = clave[0]
        asignatura = clave[1]
        paginas = por_curso[clave]
        # SIEMPRE se combina con dedupe (S11), incluso con una sola página:
        # la pantalla de revisión combina igual, y lo que el Excel escribe debe
        # ser EXACTAMENTE lo que la GUI mostró (S8). Sin duplicados, es la
        # identidad: misma hoja y mismas filas que el comportamiento histórico.
        base = dict(paginas[0])
        base["estudiantes"] = combinar_estudiantes(paginas)
        planilla_final = base

        # Nombre de hoja legible y estable: "Curso {grupo} - {asignatura corta}".
        asignatura_corta = asignatura[:22].strip()
        nombre_hoja = f"Curso {grupo} - {asignatura_corta}"[:31]
        original = nombre_hoja
        i = 2
        while nombre_hoja in nombres_usados:
            nombre_hoja = f"{original[:28]} ({i})"
            i += 1
        nombres_usados.add(nombre_hoja)

        ws = wb.create_sheet(title=nombre_hoja)
        _escribir_hoja(ws, planilla_final)

    wb.save(ruta_salida)
    return ruta_salida


if __name__ == "__main__":
    # Datos transcritos de la planilla de ejemplo (Pedro Castro Monsalvo,
    # grupo 0302, periodo 3, Educación Física) para validar el formato del Excel.
    # En producción esta lista la llena la extracción por visión, celda a celda.
    estudiantes = [
        (1, "ALFARO OCHOA LYHAM ANDRES", [45, 45], [40, 50]),
        (2, "ARIAS ECHAVEZ CELESTE SOFIA", [45, 45], [40, 45]),
        (3, "ATENCIO SIERRA KEYLLER DAVID", [45, 50], [40, 45]),
        (4, "BOTELLO ORTIZ ALVARO JOSUE", [45, 50], [45, 40]),
        (5, "CASTRO PINTO SALOMON", [45, 45], [45, 40]),
        (6, "CONTRERAS CARPIO ELIO DE JESUS", [45, 45], [45, 45]),
        (7, "DIAZ QUINTERO DYLAN DAVID", [45, 45], [45, 50]),
        (8, "ESCAMILLA GUTIERREZ ANA JULIA", [45, 45], [40, 45]),
        (9, "ESCORCIA PEDROZA MARIA SALOME", [45, 45], [40, 45]),
        (10, "GARCIA MARIN LUISA FERNANDA", [45, 44], [40, 45]),
        (11, "GOMEZ AMAYA MARIA VICTORIA", [45, 44], [45, 50]),
        (12, "GONZALEZ CASTRO LIAM DAVID", [45, 45], [40, 40]),
        (13, "GUTIERREZ ARAGON MARJALYS", [40, 40], [40, 35]),
        (14, "MARTELO ARIAS JUAN MIGUEL", [45, 50], [45, 45]),
        (15, "MARTINEZ CASTRO ELIAS DAVID", [45, 45], [40, 50]),
        (16, "MERIÑO THOMPSON DAIRO JUNIOR", [45, 45], [40, 40]),
        (17, "MOLINA OCHOA VALENTINO JOSE", [45, 45], [40, 50]),
        (18, "MURGAS MARTINEZ ERIYETH ANTONELA", [45, 50], [45, 50]),
        (19, "NAVARRO AMARIS JEREMY", [44, 40], [40, 45]),
        (20, "NAVARRO BUSTO AMANDA SOFIA", [45, 46], [40, 50]),
        (21, "PEÑA MUÑOZ ALEXANDRA", [45, 45], [40, 45]),
        (22, "PEÑA MUÑOZ MARIA FERNANDA", [45, 45], [45, 50]),
        (23, "PEREIRA JAIMES KALET DAVID", [45, 44], [40, 40]),
        (24, "PEREIRA JAIMES SEBASTIAN DAVID", [45, 45], [40, 40]),
        (25, "PEREIRA PEÑA KATHELYN VICTORIA", [45, 40], None),
        (26, "PEREIRA VILLEGAS ABRAHAM DAVID", [45, 40], [40, 40]),
        (27, "RASGO RAMIREZ JHAZIEL", [45, 50], [45, 45]),
        (28, "REBOLLEDO DE LA HOZ JORGE DE JESUS", [45, 42], [40, 40]),
        (29, "RIOS MUÑOZ MARIA FERNANDA", [45, 40], [40, 40]),
        (30, "ROMERO MACHUCA DANIELA PATRICIA", [45, 46], [40, 50]),
        (31, "SALAZAR MALDONADO IAN LUCA", [45, 45], [40, 40]),
        (32, "TERAN DE LA HOZ MARIA ELENA", [45, 45], [40, 45]),
        (33, "VENERA ARAMENDIZ ANDRES DAVID", [45, 45], [35, 40]),
        (34, "YEPES ALVAREZ VALERY", [45, 40], [40, 45]),
        (35, "ZAGARRA MEJIA JUAN DIEGO", [45, 46], [40, 50]),
    ]

    def encabezado_base(grupo):
        return {
            "institucion": "INST. ED. TEC. INDUSTRIAL PEDRO CASTRO MONSALVO",
            "sede": "SEDE CINCO DE ENERO",
            "año_lectivo": "2026",
            "jornada": "MAÑANA",
            "grupo": grupo,
            "asignatura": "0300501 EDUCACION FISICA REC. Y DEPORTES",
            "docente": "Narlis Ester Farelo Calderon",
            "periodo": 3,
        }

    def planilla_para(grupo):
        return {
            "encabezado": encabezado_base(grupo),
            "estudiantes": [
                {
                    "no": no, "nombre": nom, "ev_anteriores": ev, "area_trabajo": at, "retirado": False,
                    "revisar": [False, True] if (no == 11) else [False, False],
                }
                for no, nom, ev, at in estudiantes
            ] + [
                {"no": 36, "nombre": "LASTRE SANCHEZ ANTONELLA", "ev_anteriores": [45], "area_trabajo": None, "retirado": True}
            ],
        }

    # Simulación del caso real: un PDF con las planillas de 3 cursos distintos
    # de la misma asignatura -> un solo Excel con una hoja por curso.
    planillas_pdf = [planilla_para("0302"), planilla_para("0401"), planilla_para("0501")]

    import os
    import tempfile

    # Ruta demo portable: usa la carpeta temporal del sistema (S5), nunca una
    # ruta fija a C:\Users\JUAN.
    ruta_salida = os.path.join(
        tempfile.gettempdir(),
        "notas_educacion_fisica_periodo3.xlsx",
    )
    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    ruta = generar_excel_asignatura(planillas_pdf, ruta_salida)
    print("Generado:", ruta)
